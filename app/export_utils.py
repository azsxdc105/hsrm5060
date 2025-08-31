#!/usr/bin/env python3
"""
Data export utilities for Excel and PDF formats
"""
import os
import io
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from flask import current_app
from app.models import Claim, InsuranceCompany, User
import logging

logger = logging.getLogger(__name__)

class DataExporter:
    """Class for exporting data to various formats"""
    
    def __init__(self):
        self.setup_fonts()
    
    def setup_fonts(self):
        """Setup fonts for PDF generation"""
        try:
            # Try to register Arabic font if available
            # For production, you would need to include Arabic font files
            pass
        except Exception as e:
            logger.warning(f"Could not setup Arabic fonts: {e}")
    
    def export_claims_to_excel(self, claims: List[Claim], filename: str = None) -> str:
        """Export claims data to Excel file"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'claims_export_{timestamp}.xlsx'
        
        # Prepare data
        data = []
        for claim in claims:
            data.append({
                'رقم المطالبة': claim.id,
                'شركة التأمين': claim.insurance_company.name_ar,
                'اسم العميل': claim.client_name,
                'رقم الهوية': claim.client_national_id,
                'رقم الوثيقة': claim.policy_number or '',
                'رقم الحادث': claim.incident_number or '',
                'تاريخ الحادث': claim.incident_date.strftime('%Y-%m-%d') if claim.incident_date else '',
                'مبلغ المطالبة': float(claim.claim_amount),
                'العملة': claim.currency,
                'نوع التغطية': 'شامل' if claim.coverage_type == 'comprehensive' else 'ضد الغير',
                'المدينة': claim.city or '',
                'الحالة': self._get_status_arabic(claim.status),
                'تاريخ الإنشاء': claim.created_at.strftime('%Y-%m-%d %H:%M'),
                'أنشأها': claim.created_by.full_name,
                'تاريخ الإرسال': claim.email_sent_at.strftime('%Y-%m-%d %H:%M') if claim.email_sent_at else '',
                'العلامات': claim.tags or ''
            })
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file with styling
        upload_folder = os.path.join(current_app.root_path, 'uploads')
        os.makedirs(upload_folder, exist_ok=True)
        filepath = os.path.join(upload_folder, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='المطالبات', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['المطالبات']
            
            # Style the header
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        return filepath
    
    def export_claims_to_pdf(self, claims: List[Claim], filename: str = None) -> str:
        """Export claims data to PDF file"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'claims_export_{timestamp}.pdf'
        
        upload_folder = os.path.join(current_app.root_path, 'uploads')
        os.makedirs(upload_folder, exist_ok=True)
        filepath = os.path.join(upload_folder, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph("تقرير المطالبات", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Summary info
        summary_data = [
            ['إجمالي المطالبات:', str(len(claims))],
            ['تاريخ التقرير:', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['إجمالي المبلغ:', f"{sum(float(claim.claim_amount) for claim in claims):,.2f} ريال"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Claims table
        table_data = [['رقم المطالبة', 'العميل', 'الشركة', 'المبلغ', 'الحالة', 'تاريخ الإنشاء']]
        
        for claim in claims:
            table_data.append([
                str(claim.id),
                claim.client_name[:20] + '...' if len(claim.client_name) > 20 else claim.client_name,
                claim.insurance_company.name_ar[:15] + '...' if len(claim.insurance_company.name_ar) > 15 else claim.insurance_company.name_ar,
                f"{float(claim.claim_amount):,.0f}",
                self._get_status_arabic(claim.status),
                claim.created_at.strftime('%Y-%m-%d')
            ])
        
        # Create table
        table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        return filepath
    
    def export_companies_to_excel(self, companies: List[InsuranceCompany], filename: str = None) -> str:
        """Export insurance companies data to Excel"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'companies_export_{timestamp}.xlsx'
        
        # Prepare data
        data = []
        for company in companies:
            data.append({
                'الرقم': company.id,
                'الاسم بالعربية': company.name_ar,
                'الاسم بالإنجليزية': company.name_en,
                'البريد الرئيسي': company.claims_email_primary,
                'البريد المساعد': company.claims_email_cc or '',
                'رابط البوابة': company.policy_portal_url or '',
                'نشط': 'نعم' if company.active else 'لا',
                'عدد المطالبات': len(company.claims),
                'تاريخ الإنشاء': company.created_at.strftime('%Y-%m-%d'),
                'ملاحظات': company.notes or ''
            })
        
        df = pd.DataFrame(data)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='شركات التأمين', index=False)
            
            # Style the worksheet
            workbook = writer.book
            worksheet = writer.sheets['شركات التأمين']
            
            # Header styling
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    def _get_status_arabic(self, status: str) -> str:
        """Convert status to Arabic"""
        status_map = {
            'draft': 'مسودة',
            'ready': 'جاهز',
            'sent': 'مرسل',
            'failed': 'فشل',
            'acknowledged': 'مستلم',
            'paid': 'مدفوع'
        }
        return status_map.get(status, status)

# Global instance
data_exporter = DataExporter()

def export_claims_excel(claims: List[Claim], filename: str = None) -> str:
    """Convenience function to export claims to Excel"""
    return data_exporter.export_claims_to_excel(claims, filename)

def export_claims_pdf(claims: List[Claim], filename: str = None) -> str:
    """Convenience function to export claims to PDF"""
    return data_exporter.export_claims_to_pdf(claims, filename)

def export_companies_excel(companies: List[InsuranceCompany], filename: str = None) -> str:
    """Convenience function to export companies to Excel"""
    return data_exporter.export_companies_to_excel(companies, filename)
